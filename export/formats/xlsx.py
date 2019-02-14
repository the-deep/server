from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from utils.common import (
    get_valid_xml_string,
    parse_date,
    parse_time,
    parse_number,
)


def xstr(string):
    return get_valid_xml_string(string, escape=False)


class WorkBook:
    """
    An xlsx workbook
    """
    def __init__(self):
        self.wb = Workbook()

    def get_active_sheet(self):
        return WorkSheet(self.wb.active)

    def create_sheet(self, title):
        return WorkSheet(self.wb.create_sheet(title))

    def save(self):
        return save_virtual_workbook(self.wb)


COL_TYPES = {
    'date': 'dd-mm-yyyy',
    'time': 'HH:MM',
    'number': '',
}

TYPE_CONVERTERS = {
    'date': parse_date,
    'time': parse_time,
    'number': parse_number,
}


class WorkSheet:
    """
    A worksheet inside a workbook
    """
    def __init__(self, ws):
        self.ws = ws

    def set_title(self, title):
        self.ws.title = title
        return self

    def auto_fit_cells_in_row(self, row_id):
        row = list(self.ws.rows)[row_id - 1]
        for cell in row:
            self.ws.column_dimensions[cell.column].width =\
                max(len(cell.value), 15)

        return self

    def append(self, rows):
        [self.ws.append(row) for row in rows]
        return self

    def _set_cell_type(self, cell, col_type):
        value = cell.value
        cell.value = value and TYPE_CONVERTERS[col_type](value)
        cell.number_format = COL_TYPES[col_type]

    def set_col_types(self, col_types):
        for col_index, col_type in col_types.items():
            for cell_t in self.ws.iter_rows(
                    min_row=2,
                    min_col=col_index + 1,
                    max_col=col_index + 1,
            ):
                if len(cell_t) < 1:
                    continue
                cell = cell_t[0]
                self._set_cell_type(cell, col_type)


class RowsBuilder:
    """
    Rows builder to build rows that permute with new rows
    """
    def __init__(self, split_sheet=None, group_sheet=None, split=True):
        self.rows = [[]]
        self.group_rows = []
        self.split_sheet = split_sheet
        self.group_sheet = group_sheet
        self.split = split

    def add_value(self, value):
        val = xstr(value)
        if self.split:
            # Add single value to end of each row
            [row.append(val) for row in self.rows]
        self.group_rows.append(val)
        return self

    def add_value_list(self, value_list):
        # Add each value in the list to end of each row
        # iteratively, that is as columns
        [self.add_value(val) for val in value_list]
        return self

    def add_rows_of_values(self, rows):
        # From a list of values, for each value
        # Duplicate all rows and append that value to one set of rows

        values = [xstr(val) for val in rows]
        num = len(values)

        if num == 0:
            self.add_value('')
            return self

        if num == 1:
            self.add_value(values[0])
            return self

        if self.split:
            oldrows = self.rows[:]

            # Make a copy of old rows num times
            for i in range(1, num):
                for j, row in enumerate(oldrows):
                    self.rows.insert(i * len(oldrows) + j, row.copy())

            # Append each value to corresponding set of rows
            for i in range(0, num):
                for j in range(0, len(oldrows)):
                    self.rows[i * len(oldrows) + j].append(values[i])

        self.group_rows.append(', '.join(values))

        return self

    def add_rows_of_value_lists(self, rows, col_span=1):
        # From a list of lists, for each list
        # Duplicate all rows and append each value of
        # that list to one set of rows

        values = []
        for row in rows:
            values.append([xstr(val) for val in row])
        num = len(values)

        if num == 0:
            self.add_value_list([''] * col_span)
            return self

        if num == 1:
            self.add_value_list(values[0])
            return self

        if self.split:
            oldrows = self.rows[:]

            # Make a copy of old rows num times
            for i in range(1, num):
                for j, row in enumerate(oldrows):
                    self.rows.insert(i * len(oldrows) + j, row.copy())

            # Append each list of values to corresponding set of rows
            for i in range(0, num):
                for j in range(0, len(oldrows)):
                    for k in range(0, len(values[i])):
                        self.rows[i * len(oldrows) + j].append(values[i][k])

        # Zip to group together elements of same column,
        # Convert each zipped to list and convert overall to list as well
        for column in list(map(list, zip(*values))):
            # Make sure each column only contains unique values
            self.group_rows.append(', '.join(
                # sorted(list(dict.fromkeys(column)))
                list(OrderedDict.fromkeys(column))
            ))

        return self

    def apply(self):
        if self.split and self.split_sheet:
            self.split_sheet.append(self.rows)

        if self.group_sheet:
            self.group_sheet.append([self.group_rows])
