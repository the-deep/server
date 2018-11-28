from collections import OrderedDict

from django.core.files.base import ContentFile

from export.models import Export
from export.formats.xlsx import WorkBook, RowsBuilder
from openpyxl.styles import Alignment
from export.mime_types import EXCEL_MIME_TYPE
from utils.common import format_date, generate_filename

import logging

logger = logging.getLogger('django')


class ExcelExporter:
    def __init__(self, decoupled=True):
        self.wb = WorkBook()

        # Create two worksheets
        if decoupled:
            self.split = self.wb.get_active_sheet()\
                .set_title('Split Assessments')
            self.group = self.wb.create_sheet('Grouped Assessments')
        else:
            self.split = None
            self.group = self.wb.get_active_sheet().set_title('Assessments')
        self.decoupled = decoupled

        # Cells to be merged
        self.merge_cells = {}
        # Initial titles
        self.titles = [
            'Date of Lead Publication',
            'Imported By',
            'Lead Title',
            'Source',
        ]
        self.col_types = {
            0: 'date',
        }
        self._titles_dict = {k: True for k in self.titles}

    def to_flattened_key_vals(self, dictdata, parents=[]):
        """
        Convert nested dictionary data to flat dict with keys and nondict
        values.
        @dictdata: nested dict structrure to be flattened
        @parents: parent titles for given level of title.
                Order [immediate parent, next parent, ... root]
        """
        flat = OrderedDict()  # to preserve order of titles
        for k, v in dictdata.items():
            if isinstance(v, dict):
                # NOTE: this might override the keys in recursive calls
                flat.update(self.to_flattened_key_vals(v, [k, *parents]))
            elif isinstance(v, list):
                # check if list elements are dict or not
                for i in v:
                    if isinstance(i, dict):
                        flat.update(
                            self.to_flattened_key_vals(i, [k, *parents])
                        )
                    else:
                        vals = flat.get(k, {}).get('value', [])
                        vals.append(i)
                        # FIXME: assigning parents is repeated every step
                        flat[k] = {'value': vals, 'parents': parents}
            else:
                # Just add key value
                flat[k] = {'value': v, 'parents': parents}
        return flat

    def add_assessments(self, assessments):
        for a in assessments:
            self.add_assessment(a)
        return self

    def add_assessment(self, assessment):
        jsondata = assessment.to_exportable_json()
        flat = self.to_flattened_key_vals(jsondata)
        rows = RowsBuilder(self.split, self.group, split=False)
        rows.add_value_list([
            format_date(assessment.lead.created_at),
            assessment.lead.created_by.username,
            assessment.lead.title,
            assessment.lead.source
        ])
        # update the titles
        for k, v in flat.items():
            if not self._titles_dict.get(k):
                self.titles.append(k)
                self._titles_dict[k] = True

        # create headers
        self.merge_cells = {}  # to store which cells to merge
        headers_dict = {}
        title_headers = []
        for i, t in enumerate(self.titles):
            v = flat.get(t)
            if v:
                v = flat[t]['value']
                val = ', '.join(v) if isinstance(v, list) else str(v)
                rows.add_value(val)
                header = flat[t]['parents'][-1]
                if not headers_dict.get(header):
                    self.merge_cells[header] = {'start': i, 'end': i}
                    title_headers.append(header.upper())
                    headers_dict[header] = True
                else:
                    self.merge_cells[header]['end'] += 1
                    title_headers.append("")
            else:
                title_headers.append("")

        headerrows = RowsBuilder(self.split, self.group, split=False)
        headerrows.add_value_list(title_headers)
        headerrows.apply()
        self.group.append([self.titles])

        if self.decoupled and self.split:
            self.split.auto_fit_cells_in_row(1)
        self.group.auto_fit_cells_in_row(1)

        rows.apply()
        return self

    def export(self, export_entity):
        # merge cells
        if self.merge_cells:
            sheet = self.wb.wb.active
            for k, v in self.merge_cells.items():
                sheet.merge_cells(
                    start_row=1, start_column=v['start'] + 1,
                    end_row=1, end_column=v['end'] + 1
                )
                cell = sheet.cell(row=1, column=v['start'] + 1)
                cell.alignment = Alignment(horizontal='center')

        self.group.set_col_types(self.col_types)
        if self.split:
            self.split.set_col_types(self.col_types)

        buffer = self.wb.save()
        filename = generate_filename('Assessments Export', 'xlsx')

        export_entity.title = filename
        export_entity.type = Export.ASSESSMENTS
        export_entity.format = 'xlsx'
        export_entity.pending = False
        export_entity.mime_type = EXCEL_MIME_TYPE

        export_entity.file.save(filename, ContentFile(buffer))
        export_entity.save()
