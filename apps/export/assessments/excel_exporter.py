from collections import OrderedDict

from django.core.files.base import ContentFile

from export.models import Export
from export.formats.xlsx import WorkBook, RowsBuilder
from openpyxl.styles import Alignment, Font
from export.mime_types import EXCEL_MIME_TYPE
from utils.common import format_date, generate_filename, underscore_to_title

import logging

logger = logging.getLogger('django')


class ExcelExporter:
    def __init__(self, decoupled=True):
        self.wb = WorkBook()

        # Create two worksheets
        if decoupled:
            self.split = self.wb.get_active_sheet()\
                .set_title('Split Assessments')
            self.group = self.wb.create_sheet('Grouped Assessments')
        else:
            self.split = None
            self.group = self.wb.get_active_sheet().set_title('Assessments')
        self.decoupled = decoupled

        # Cells to be merged
        self.merge_cells = {}
        # Initial titles
        self.lead_titles = [
            'Date of Lead Publication',
            'Imported By',
            'Lead Title',
            'Source',
        ]
        self.titles = [*self.lead_titles]
        self.col_types = {
            0: 'date',
        }
        self._titles_dict = {k: True for k in self.titles}

        self._headers_titles = OrderedDict()
        self._headers_added = False
        self._excel_rows = []
        self._title_headers = []
        self._headers_dict = {}
        self._flats = []
        self._assessments = []

    def to_flattened_key_vals(self, dictdata, parents=[]):
        """
        Convert nested dictionary data to flat dict with keys and nondict
        values.
        @dictdata: nested dict structrure to be flattened
        @parents: parent titles for given level of title.
                Order [immediate parent, next parent, ... root]
        """
        flat = OrderedDict()  # to preserve order of titles
        for k, v in dictdata.items():
            if isinstance(v, dict):
                # NOTE: this might override the keys in recursive calls
                flat.update(self.to_flattened_key_vals(v, [k, *parents]))
            elif isinstance(v, list):
                # check if list elements are dict or not
                for i in v:
                    if isinstance(i, dict):
                        flat.update(
                            self.to_flattened_key_vals(i, [k, *parents])
                        )
                    else:
                        vals = flat.get(k, {}).get('value', [])
                        vals.append(i)
                        # FIXME: assigning parents is repeated every step
                        flat[k] = {'value': vals, 'parents': parents}
            else:
                # Just add key value
                flat[k] = {'value': v, 'parents': parents}
        return flat

    def add_assessments(self, assessments):
        for a in assessments:
            self.add_assessment(a)
        return self

    def add_assessment(self, assessment):
        jsondata = assessment.to_exportable_json()
        flat = self.to_flattened_key_vals(jsondata)
        self._flats.append(flat)
        self._assessments.append(assessment)

        # update the titles
        for k, v in flat.items():
            parent = v['parents'][-1]
            header_titles = self._headers_titles.get(parent, [])

            if k not in header_titles:
                header_titles.append(k)
            self._headers_titles[parent] = header_titles
            '''
            if not self._titles_dict.get(k):
                self.titles.append(k)
                self._titles_dict[k] = True
            '''
        return self

    def get_titles(self):
        return [
            *self.lead_titles,
            *[y for k, v in self._headers_titles.items() for y in v]
        ]

    def assessments_to_rows(self):
        for index, assessment in enumerate(self._assessments):
            rows = RowsBuilder(self.split, self.group, split=False)
            rows.add_value_list([
                format_date(assessment.lead.created_at),
                assessment.lead.created_by.username,
                assessment.lead.title,
                assessment.lead.source
            ])
            headers_dict = {}
            flat = self._flats[index]
            for i, t in enumerate(self.get_titles()):
                v = flat.get(t)
                if not v and t not in self.lead_titles:
                    rows.add_value("")
                    self._title_headers.append("")
                    continue
                elif not v:
                    self._title_headers.append("")
                    continue

                v = flat[t]['value']
                val = ', '.join([str(x) for x in v]) if isinstance(v, list) else str(v)
                rows.add_value(val)
                header = flat[t]['parents'][-1]

                if not self._headers_dict.get(header):
                    self._title_headers.append(header.upper())
                    self._headers_dict[header] = True
                else:
                    self.merge_cells[header]['end'] += 1

                if not headers_dict.get(header):
                    self.merge_cells[header] = {'start': i, 'end': i}
                    headers_dict[header] = True
                else:
                    self._title_headers.append("")

            self._excel_rows.append(rows)

    def export(self):
        # Generate rows
        self.assessments_to_rows()

        # add header rows
        headerrows = RowsBuilder(self.split, self.group, split=False)
        headerrows.add_value_list(self._title_headers)
        headerrows.apply()

        self.group.append([self.get_titles()])

        if self.decoupled and self.split:
            self.split.auto_fit_cells_in_row(1)
        self.group.auto_fit_cells_in_row(1)

        # add rows
        for rows in self._excel_rows:
            rows.apply()

        # merge cells
        if self.merge_cells:
            sheet = self.wb.wb.active
            for k, v in self.merge_cells.items():
                sheet.merge_cells(
                    start_row=1, start_column=v['start'] + 1,
                    end_row=1, end_column=v['end'] + 1
                )
                cell = sheet.cell(row=1, column=v['start'] + 1)
                cell.alignment = Alignment(horizontal='center')

        self.group.set_col_types(self.col_types)
        if self.split:
            self.split.set_col_types(self.col_types)

        buffer = self.wb.save()
        filename = generate_filename('Assessments Export', 'xlsx')
        return filename, Export.XLSX, EXCEL_MIME_TYPE, ContentFile(buffer)


class NewExcelExporter:
    def __init__(self, sheets_data):
        """
        sheets_data = {
            sheet1: {
                grouped_col: [
                    { col1: val1, col2: val2, col3: val3 },
                    { col1: val1, col2: val2, col3: val3 },
                    ...
                ],
                ungrouped_col: [
                    val1,
                    val2,
                    ...
                ],
            },
            sheet2: {
                ...
            }
        }
        """
        self.wb = WorkBook()
        self.sheets_data = sheets_data
        self.wb_sheets = {}
        self.sheet_headers = {}

        # TODO: validate all data in each col, to see if each of them have same
        # structure(keys)
        for sheet, sheet_data in sheets_data.items():
            self.wb_sheets[sheet] = self.wb.create_sheet(underscore_to_title(sheet))
            self.sheet_headers[sheet] = {}
            for col, data in sheet_data.items():
                if data and isinstance(data[0], dict):
                    self.sheet_headers[sheet][col] = data[0].keys()
                else:
                    self.sheet_headers[sheet][col] = []

    def add_headers(self):
        for sheet, headerinfo in self.sheet_headers.items():
            header_row = []
            sub_header_row = []
            for header, info in headerinfo.items():
                header_row.append(underscore_to_title(str(header)))
                # Also add empty cells to account for sub headers
                if info:
                    header_row.extend([""] * (len(info) - 1))
                    sub_header_row.extend([underscore_to_title(x) for x in info])
                else:
                    sub_header_row.append("")

            # Append header rows to sheet
            self.wb_sheets[sheet].append([header_row, sub_header_row])
            # Merge/style Headers
            counter = 1
            for header, info in headerinfo.items():
                wb_sheet = self.wb_sheets[sheet].ws
                if info:
                    wb_sheet.merge_cells(
                        start_row=1,
                        start_column=counter,
                        end_row=1,
                        end_column=counter + len(info) - 1
                    )
                    counter += len(info)
                else:
                    counter += 1
                # Styling
                cell = wb_sheet.cell(row=1, column=counter)
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
            # Style sub headers
            for i, header in enumerate(sub_header_row):
                cell = wb_sheet.cell(row=2, column=i + 1)
                cell.font = Font(bold=True)
            self.wb_sheets[sheet].auto_fit_cells_in_row(1)
            self.wb_sheets[sheet].auto_fit_cells_in_row(2)

    def add_data_rows_to_sheets(self):
        for sheet, sheet_data in self.sheets_data.items():
            zipped = zip(*sheet_data.values())
            for row_data in zipped:
                row = []
                for each in row_data:
                    rowdata = each.values() if isinstance(each, dict) else [each]
                    row.extend(rowdata)
                self.wb_sheets[sheet].append([row])

    def export(self):
        # Write cols header first
        self.add_headers()

        # Add data rows
        self.add_data_rows_to_sheets()

        # Remove default sheet
        self.wb.wb.remove(self.wb.wb.get_sheet_by_name('Sheet'))

        buffer = self.wb.save()
        filename = generate_filename('Assessments Export', 'xlsx')
        return filename, Export.XLSX, EXCEL_MIME_TYPE, ContentFile(buffer)
