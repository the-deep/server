import logging
from openpyxl import load_workbook

from ..models import Sheet, Field
from datetime import datetime

from utils.common import (
    LogTime,
    excel_to_python_date_format,
    format_date_or_iso,
)

logger = logging.getLogger(__name__)


def get_cell_value(row, column):
    try:
        return row[column].value
    except IndexError:
        return None


def is_row_empty(row, columns):
    for column in columns:
        if column is not None:
            value = get_cell_value(row, column)
            if value is not None:
                return False
    return True


def get_excel_value(cell):
    value = cell.value
    if value is not None and isinstance(value, datetime):
        dateformat = cell.number_format
        # try casting to python format
        python_format = excel_to_python_date_format(
            dateformat
        )
        return format_date_or_iso(
            cell.value, python_format
        )
    elif value is not None and not isinstance(value, str):
        return str(cell.internal_value)
    return str(value)


@LogTime()
def extract(book):
    options = book.options if book.options else {}
    Sheet.objects.filter(book=book).delete()  # Delete all previous sheets
    with book.get_file() as xlsx_file:
        workbook = load_workbook(xlsx_file, data_only=True, read_only=True)
        for sheet_key, wb_sheet in enumerate(workbook.worksheets):
            sheet_options = options.get('sheets', {}).get(str(sheet_key), {})
            if sheet_options.get('skip', False):
                continue
            sheet = Sheet.objects.create(
                title=wb_sheet.title,
                book=book,
            )

            sheet_rows = []

            no_headers = sheet_options.get('no_headers', False)

            max_col_length = 1
            for row in wb_sheet.iter_rows():
                row_data = get_row_data(row)

                length = len(row_data)
                if length > max_col_length:
                    max_col_length = length

                sheet_rows.append(row_data)

            if not sheet_rows:
                return

            if no_headers:
                fields = [
                    Field(title=f'Column {x}', sheet=sheet, ordering=x, data=[])
                    for x in range(max_col_length)
                ]
            else:
                fields = []
                for x in range(max_col_length):
                    row_len = len(sheet_rows[0])
                    title_val = sheet_rows[0][x]['value'] if row_len > x else None
                    title = title_val or f'Column {x}'
                    fields.append(Field(title=title, sheet=sheet, ordering=x, data=[]))

            empty_value = {
                'value': None,
                'invalid': False,
                'empty': True
            }
            # Now append data to fields
            for row in sheet_rows:
                row_len = len(row)
                for x in range(max_col_length):
                    cell_data = row[x] if x < row_len else {**empty_value}
                    fields[x].data.append(cell_data)

            # Bulk save fields
            Field.objects.bulk_create(fields)

            sheet.data_row_index = 0 if no_headers else 1
            sheet.save()


def get_row_data(row):
    """
    Returns cells values for the row.
    """
    data = []
    max_data_col = 0  # max column number containing data
    curr_col = 0
    for cell in row:
        if cell.value is not None:
            max_data_col = curr_col
        value = get_excel_value(cell)
        data.append({
            'value': value,
            'empty': value is None,
            'invalid': False
        })
        curr_col += 1
    # Now clip the data beyond which there is nothing
    return data[:max_data_col + 1]
