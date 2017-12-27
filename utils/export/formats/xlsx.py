from openpyxl import Workbook
# from openpyxl.writer.excel import save_virtual_workbook

from utils.export.common import get_valid_xml_string as xstr


class WorkBook:
    def __init__(self):
        self.wb = Workbook()

    def get_active_sheet(self):
        return WorkSheet(self.wb.active)

    def create_sheet(self, title):
        return WorkSheet(self.wb.create_sheet(title))

    def save_to(self, filename):
        self.wb.save(filename)


class WorkSheet:
    def __init__(self, ws):
        self.ws = ws

    def set_title(self, title):
        self.ws.title = title
        return self

    def auto_fit_cells_in_row(self, row_id):
        row = list(self.ws.rows)[row_id - 1]
        for cell in row:
            self.ws.column_dimensions[cell.column].width =\
                max(len(cell.value), 15)

        return self

    def append(self, rows):
        [self.ws.append(row) for row in rows]

        return self


class RowsBuilder:
    def __init__(self, split_sheet=None, group_sheet=None):
        self.rows = [[]]
        self.group_rows = []
        self.split_sheet = split_sheet
        self.group_sheet = group_sheet

    def add_value(self, value):
        val = xstr(value)
        [row.append(val) for row in self.rows]
        self.group_rows.append(val)
        return self

    def add_value_list(self, value_list):
        [self.add_value(val) for val in value_list]
        return self

    def add_rows_of_values(self, rows):
        values = [xstr(val) for val in rows]
        num = len(values)

        if num == 0:
            self.add_value('')
            return self

        if num == 1:
            self.add_value(values[0])
            return self

        oldrows = self.rows[:]
        for i in range(1, num):
            for j, row in enumerate(oldrows):
                self.rows.insert(i * len(oldrows) + j, row.copy())

        for i in range(0, num):
            for j in range(0, len(oldrows)):
                self.rows[i * len(oldrows) + j].append(values[i])

        self.group_rows.append(', '.join(values))

        return self

    def add_rows_of_value_lists(self, rows):
        values = []
        for row in rows:
            values.append([xstr(val) for val in row])
        num = len(values)

        if num == 0:
            self.add_value('')

        if num == 1:
            self.add_value_list(values[0])

        oldrows = self.rows[:]
        for i in range(1, num):
            for j, row in enumerate(oldrows):
                self.rows.insert(i * len(oldrows) + j, row.copy())

        for i in range(0, num):
            for j in range(0, len(oldrows)):
                for k in range(0, len(values[i])):
                    self.rows[i * len(oldrows) + j].append(values[i][k])

        for column in list(map(list, zip(*values))):
            self.group_rows.append(', '.join(list(dict.fromkeys(column))))

        return self

    def apply(self):
        if self.split_sheet:
            self.split_sheet.append(self.rows)

        if self.group_sheet:
            self.group_sheet.append(self.group_rows)
