from openpyxl import load_workbook

from ..models import Sheet, Field


def get_cell_value(row, column):
    try:
        return row[column].value
    except IndexError:
        return None


def is_row_empty(row, columns):
    for column in columns:
        if column is not None:
            value = get_cell_value(row, column)
            if value is not None:
                return False
    return True


def extract(book):
    options = book.options if book.options else {}
    Sheet.objects.filter(book=book).delete()  # Delete all previous sheets
    with book.get_file() as xlsx_file:
        workbook = load_workbook(xlsx_file, data_only=True, read_only=True)
        for sheet_key, wb_sheet in enumerate(workbook.worksheets):
            sheet_options = options.get('sheets', {}).get(str(sheet_key), {})
            if sheet_options.get('skip', False):
                continue
            sheet = Sheet.objects.create(
                title=wb_sheet.title,
                book=book,
            )
            header_index = sheet_options.get('header_row', 1)
            no_headers = sheet_options.get('no_headers', False)
            data_index = sheet_options.get('data_row_index', header_index + 1)

            if no_headers:
                data_index -= 1

            # Fields
            header_row = list(
                wb_sheet.iter_rows(
                    min_row=header_index, max_row=header_index + 1,
                )
            )[0]
            fields = []
            columns = []
            ordering = 1
            for cell in header_row:
                if cell.value is not None:
                    columns.append(cell.column)
                    fields.append(
                        Field(
                            title=(cell.value if not no_headers
                                   else 'Column ' + str(ordering)),
                            sheet=sheet,
                            ordering=ordering,
                        )
                    )
                else:
                    fields.append(None)
                ordering += 1

            Field.objects.bulk_create(
                [field for field in fields if field is not None]
            )

            fields_data = {}
            # Data
            for _row in wb_sheet.iter_rows(min_row=data_index):
                if is_row_empty(_row, columns):
                    continue
                try:
                    for index, field in enumerate(fields):
                        if field is None:
                            continue
                        value = _row[index].value

                        if value is not None and not isinstance(value, str):
                            value = _row[index].internal_value

                        field_data = fields_data.get(field.id, [])
                        field_data.append({
                            'value': value,
                            'empty': False,
                            'invalid': False
                        })
                        fields_data[field.id] = field_data

                except Exception:
                    pass

            # Save field
            for field in sheet.fields:
                field.data = fields_data.get(field.id, [])
                field.save()
